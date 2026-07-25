
import openpyxl, os, io, sys, re
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
DB = r'd:\DoAnTotNghiepCuNhan\HeThongQuanLyKienTap\DB'

# Read all xlsx files
def load(name):
    wb = openpyxl.load_workbook(os.path.join(DB, name), read_only=True, data_only=True)
    return wb

# Collect unique entities
students = {}  # mssv -> {ho_ten, lop, gvhd}
factories = set()
teachers = set()
visits = []  # (mssv, factory, date, form, score, bonus)

# File 1: Diem qua trinh - main data source
wb = load(u'D\u1eef li\u1ec7u \u0111i\u1ec3m qu\u00e1 tr\u00ecnh TQNM 2024-2025.xlsx')
ws = wb['Sheet1']
rows = list(ws.iter_rows(values_only=True))
wb.close()

for r in rows[6:]:  # skip header rows (1-6)
    if r[0] is None or not isinstance(r[0], (int, float)):
        continue
    mssv = str(r[1]).strip().strip("'") if r[1] else None
    if not mssv: continue
    ho = str(r[2]).strip() if r[2] else ''
    ten = str(r[3]).strip() if r[3] else ''
    ho_ten = f"{ho} {ten}".strip()
    lop = str(r[4]).strip() if r[4] else ''
    students[mssv] = {'ho_ten': ho_ten, 'lop': lop}
    
    # Each row has up to 7 factory visits at columns 6,11,16,21,26,31,36
    for i in range(7):
        base = 6 + i*5
        if base >= len(r): break
        nm = r[base]
        if not nm or nm == ' ': continue
        nm = str(nm).strip()
        factories.add(nm)
        dt = r[base+1] if base+1 < len(r) else None
        form = r[base+2] if base+2 < len(r) else None
        score = r[base+3] if base+3 < len(r) else None
        bonus = r[base+4] if base+4 < len(r) else None
        
        date_str = None
        if isinstance(dt, datetime):
            date_str = dt.strftime('%Y-%m-%d')
        elif isinstance(dt, str) and dt.strip():
            date_str = dt.strip()
        
        form_str = 'TrucTiep'
        if form and str(form).strip().lower() in ('online',):
            form_str = 'TrucTuyen'
        elif form and 'do' in str(form).strip().lower():
            form_str = 'TrucTiep'  # Tu do is still offline
            
        sc = None
        if score and score != ' ':
            try: sc = float(score)
            except: pass
        
        bn = None
        if bonus and bonus != ' ':
            try: bn = float(bonus)
            except: pass
            
        visits.append({'mssv': mssv, 'factory': nm, 'date': date_str, 'form': form_str, 'score': sc, 'bonus': bn})

# File 2: Kien tap 24-25 - get GVHD info
wb2 = load(u'D\u1eef li\u1ec7u Kien tap 24-25 (G\u1eedi SV ki\u1ec3m tra).xlsx')
for sn in wb2.sheetnames:
    if sn == 'Sheet1': continue
    ws2 = wb2[sn]
    rows2 = list(ws2.iter_rows(values_only=True))
    for r in rows2:
        if r[0] is None or not isinstance(r[0], (int, float)): continue
        mssv = str(r[1]).strip() if r[1] else None
        if not mssv: continue
        gvhd = str(r[6]).strip() if len(r) > 6 and r[6] else None
        if mssv in students and gvhd:
            students[mssv]['gvhd'] = gvhd
            teachers.add(gvhd)
wb2.close()

# File 3 & 4: DS Acecook & Heineken - get GVHD
for fname, code in [('DS TQNM ACECOOK T5-2025.xlsx','ACE'), ('DSSV TQNM Heineken 22052025.xlsx','HNK')]:
    wb3 = load(fname)
    for sn in wb3.sheetnames:
        ws3 = wb3[sn]
        for r in ws3.iter_rows(values_only=True):
            if r[0] is None or not isinstance(r[0], (int, float)): continue
            mssv = str(r[1]).strip() if r[1] else None
            if not mssv: continue
            ho = str(r[2]).strip() if r[2] else ''
            ten = str(r[3]).strip() if r[3] else ''
            ht = f"{ho} {ten}".strip()
            lop = str(r[4]).strip() if r[4] else ''
            gvhd = str(r[8]).strip() if len(r) > 8 and r[8] else None
            if mssv not in students:
                students[mssv] = {'ho_ten': ht, 'lop': lop}
            if gvhd:
                students[mssv]['gvhd'] = gvhd
                teachers.add(gvhd)
    wb3.close()

# Ensure all teachers collected
for s in students.values():
    if 'gvhd' in s and s['gvhd']:
        teachers.add(s['gvhd'])

# Extract khoa from lop
def extract_khoa(lop):
    m = re.match(r'(\d+)', lop)
    if m:
        return m.group(1)
    return None

khoa_set = set()
for s in students.values():
    k = extract_khoa(s.get('lop',''))
    if k: khoa_set.add(k)

# Factory name normalization
factory_map = {
    'Acecook': 'Công ty CP Acecook Việt Nam',
    'Acccook': 'Công ty CP Acecook Việt Nam',
    'Yakult': 'Công ty TNHH Yakult Việt Nam',
    'Kewpie': 'Công ty TNHH Kewpie Việt Nam',
    'Heineken': 'Nhà máy Bia Heineken Vũng Tàu',
    'Satori': 'Công ty TNHH Satori Việt Nam',
    'Ajinomoto Long Thành': 'Công ty Ajinomoto Việt Nam - CN Long Thành',
    'Ajinomoto Biên Hòa': 'Công ty Ajinomoto Việt Nam - CN Biên Hòa',
    'Aeon': 'Công ty TNHH Aeon Việt Nam',
    'Nutifood': 'Công ty CP Nutifood',
    'TTC': 'Công ty CP TTC',
}

def esc(s):
    if s is None: return 'NULL'
    return "N'" + str(s).replace("'", "''") + "'"

# Generate SQL
out = []
out.append("/* ============================================================================")
out.append("   SCRIPT IMPORT DỮ LIỆU - HỆ THỐNG QUẢN LÝ KIẾN TẬP")
out.append("   Tự động sinh từ các file xlsx nguồn")
out.append("   Chạy SAU khi đã tạo database bằng QLKienTap_Database.sql")
out.append("   ============================================================================ */")
out.append("")
out.append("USE QLKienTap;")
out.append("GO")
out.append("")

# 1. NamHoc
out.append("-- ============ 1. NamHoc ============")
out.append("SET IDENTITY_INSERT NamHoc ON;")
out.append("INSERT INTO NamHoc (id, ten_nam_hoc, ngay_bat_dau, ngay_ket_thuc) VALUES")
out.append("  (1, N'2024-2025', '2024-09-01', '2025-08-31');")
out.append("SET IDENTITY_INSERT NamHoc OFF;")
out.append("GO")
out.append("")

# 2. HocKy
out.append("-- ============ 2. HocKy ============")
out.append("SET IDENTITY_INSERT HocKy ON;")
out.append("INSERT INTO HocKy (id, nam_hoc_id, ten_hoc_ky, ngay_bat_dau, ngay_ket_thuc) VALUES")
out.append("  (1, 1, N'Học kỳ 1', '2024-09-01', '2025-01-31'),")
out.append("  (2, 1, N'Học kỳ 2', '2025-02-01', '2025-06-30');")
out.append("SET IDENTITY_INSERT HocKy OFF;")
out.append("GO")
out.append("")

# 3. Khoa
out.append("-- ============ 3. Khoa ============")
out.append("SET IDENTITY_INSERT Khoa ON;")
sorted_khoa = sorted(khoa_set, key=lambda x: int(x))
khoa_id_map = {}
for i, k in enumerate(sorted_khoa, 1):
    khoa_id_map[k] = i
    prefix = k
    nam = 2000 + int(k) if int(k) < 50 else 1900 + int(k)
    ten = f"{k}\u0110HTP"
    out.append(f"INSERT INTO Khoa (id, ten_khoa, nam_nhap_hoc) VALUES ({i}, N'{ten}', {nam});")
out.append("SET IDENTITY_INSERT Khoa OFF;")
out.append("GO")
out.append("")

# 4. TaiKhoan - generate for all teachers and students
out.append("-- ============ 4. TaiKhoan ============")
out.append("SET IDENTITY_INSERT TaiKhoan ON;")
tk_id = 1
teacher_tk = {}
student_tk = {}

# Teachers first
sorted_teachers = sorted(teachers)
for t in sorted_teachers:
    uname = t.lower().replace(' ', '.')
    # Remove diacritics roughly
    out.append(f"INSERT INTO TaiKhoan (id, ten_dang_nhap, mat_khau_hash, vai_tro, trang_thai, phai_doi_mat_khau)")
    out.append(f"  VALUES ({tk_id}, N'gv{tk_id:03d}', N'$2b$10$hash_placeholder', N'GiangVien', N'HoatDong', 0);")
    teacher_tk[t] = tk_id
    tk_id += 1

# Admin account
out.append(f"INSERT INTO TaiKhoan (id, ten_dang_nhap, mat_khau_hash, vai_tro, trang_thai, phai_doi_mat_khau)")
out.append(f"  VALUES ({tk_id}, N'admin01', N'$2b$10$hash_placeholder', N'QuanLyKhoa', N'HoatDong', 0);")
admin_tk = tk_id
tk_id += 1

# Students
sorted_students = sorted(students.keys())
for mssv in sorted_students:
    out.append(f"INSERT INTO TaiKhoan (id, ten_dang_nhap, mat_khau_hash, vai_tro, trang_thai, phai_doi_mat_khau)")
    out.append(f"  VALUES ({tk_id}, N'{mssv}', N'$2b$10$hash_placeholder', N'SinhVien', N'HoatDong', 1);")
    student_tk[mssv] = tk_id
    tk_id += 1

out.append("SET IDENTITY_INSERT TaiKhoan OFF;")
out.append("GO")
out.append("")

# 5. GiangVien
out.append("-- ============ 5. GiangVien ============")
out.append("SET IDENTITY_INSERT GiangVien ON;")
gv_id_map = {}
gv_id = 1
for t in sorted_teachers:
    gv_id_map[t] = gv_id
    out.append(f"INSERT INTO GiangVien (id, ma_gv, ho_ten, taikhoan_id, du_dk_hoi_dong)")
    out.append(f"  VALUES ({gv_id}, N'GV{gv_id:03d}', {esc(t)}, {teacher_tk[t]}, 1);")
    gv_id += 1
out.append("SET IDENTITY_INSERT GiangVien OFF;")
out.append("GO")
out.append("")

# 6. SinhVien
out.append("-- ============ 6. SinhVien ============")
out.append("SET IDENTITY_INSERT SinhVien ON;")
sv_id_map = {}
sv_id = 1
for mssv in sorted_students:
    s = students[mssv]
    lop = s.get('lop', '')
    khoa_num = extract_khoa(lop)
    kid = khoa_id_map.get(khoa_num, 1) if khoa_num else 1
    # Determine hoc_lai
    hoc_lai = 0
    if khoa_num and int(khoa_num) < 13:
        hoc_lai = 1
    sv_id_map[mssv] = sv_id
    out.append(f"INSERT INTO SinhVien (id, mssv, ho_ten, taikhoan_id, khoa_id, ten_lop, hoc_lai)")
    out.append(f"  VALUES ({sv_id}, N'{mssv}', {esc(s['ho_ten'])}, {student_tk[mssv]}, {kid}, {esc(lop)}, {hoc_lai});")
    sv_id += 1
out.append("SET IDENTITY_INSERT SinhVien OFF;")
out.append("GO")
out.append("")

# 7. NhaMay
out.append("-- ============ 7. NhaMay ============")
out.append("SET IDENTITY_INSERT NhaMay ON;")
nm_id_map = {}
nm_id = 1
nhom_nganh_map = {
    'Acecook': 'Mì ăn liền',
    'Acccook': 'Mì ăn liền',
    'Yakult': 'Sữa lên men',
    'Kewpie': 'Sốt - Gia vị',
    'Heineken': 'Đồ uống',
    'Satori': 'Nước giải khát',
    'Ajinomoto Long Thành': 'Gia vị - Thực phẩm chế biến',
    'Ajinomoto Biên Hòa': 'Gia vị - Thực phẩm chế biến',
    'Aeon': 'Bán lẻ - Siêu thị',
    'Nutifood': 'Sữa - Dinh dưỡng',
    'TTC': 'Đường - Nông sản',
}
online_map = {'Ajinomoto Long Thành': 1, 'Ajinomoto Biên Hòa': 1}

for f in sorted(factories):
    ten = factory_map.get(f, f)
    nhom = nhom_nganh_map.get(f, None)
    online = online_map.get(f, 0)
    nm_id_map[f] = nm_id
    out.append(f"INSERT INTO NhaMay (id, ten_nha_may, nhom_nganh, ho_tro_truc_tiep, ho_tro_truc_tuyen)")
    out.append(f"  VALUES ({nm_id}, {esc(ten)}, {esc(nhom)}, 1, {online});")
    nm_id += 1
# Map Acccook to same as Acecook
if 'Acccook' in nm_id_map and 'Acecook' in nm_id_map:
    nm_id_map['Acccook'] = nm_id_map['Acecook']
out.append("SET IDENTITY_INSERT NhaMay OFF;")
out.append("GO")
out.append("")

# 8. DotKienTap
out.append("-- ============ 8. DotKienTap ============")
out.append("SET IDENTITY_INSERT DotKienTap ON;")
out.append("INSERT INTO DotKienTap (id, ten_dot, nam_hoc_id, hoc_ky_id, ngay_bat_dau, ngay_ket_thuc, trang_thai)")
out.append("  VALUES (1, N'Đợt kiến tập 2024-2025', 1, 2, '2024-03-01', '2025-06-30', N'DaKetThuc');")
out.append("SET IDENTITY_INSERT DotKienTap OFF;")
out.append("GO")
out.append("")

# 9. LichKienTap
out.append("-- ============ 9. LichKienTap ============")
out.append("SET IDENTITY_INSERT LichKienTap ON;")
out.append("INSERT INTO LichKienTap (id, dot_kien_tap_id, khoa_id, ten_lich,")
out.append("  tg_mo_dang_ky_tu, tg_mo_dang_ky_den, tg_dien_ra_tu, tg_dien_ra_den,")
out.append("  han_chot_nop_bao_cao, han_chot_diem, trang_thai)")
out.append("  VALUES (1, 1, " + str(khoa_id_map.get('13', 1)) + ", N'Lịch kiến tập Khóa 13 - NH 2024-2025',")
out.append("  '2024-03-01', '2024-03-15', '2024-03-16', '2025-06-15',")
out.append("  '2025-06-30', '2025-07-15', N'DaKetThuc');")
out.append("SET IDENTITY_INSERT LichKienTap OFF;")
out.append("GO")
out.append("")

# 10. LichKienTap_SinhVien
out.append("-- ============ 10. LichKienTap_SinhVien ============")
out.append("SET IDENTITY_INSERT LichKienTap_SinhVien ON;")
lksv_id = 1
lksv_map = {}
for mssv in sorted_students:
    if mssv in sv_id_map:
        s = students[mssv]
        hoc_lai_flag = 0
        khoa_num = extract_khoa(s.get('lop',''))
        if khoa_num and int(khoa_num) < 13:
            hoc_lai_flag = 1
        lan = 2 if hoc_lai_flag else 1
        out.append(f"INSERT INTO LichKienTap_SinhVien (id, lich_kien_tap_id, sinh_vien_id, lan_dang_ky, trang_thai)")
        out.append(f"  VALUES ({lksv_id}, 1, {sv_id_map[mssv]}, {lan}, N'DangThucHien');")
        lksv_map[mssv] = lksv_id
        lksv_id += 1
out.append("SET IDENTITY_INSERT LichKienTap_SinhVien OFF;")
out.append("GO")
out.append("")

# 11. PhanCongGVHD
out.append("-- ============ 11. PhanCongGVHD ============")
out.append("SET IDENTITY_INSERT PhanCongGVHD ON;")
pc_id = 1
for mssv in sorted_students:
    s = students[mssv]
    gvhd = s.get('gvhd')
    if gvhd and gvhd in gv_id_map and mssv in lksv_map:
        out.append(f"INSERT INTO PhanCongGVHD (id, lich_kien_tap_sinh_vien_id, giang_vien_id, trang_thai)")
        out.append(f"  VALUES ({pc_id}, {lksv_map[mssv]}, {gv_id_map[gvhd]}, N'DangHoatDong');")
        pc_id += 1
out.append("SET IDENTITY_INSERT PhanCongGVHD OFF;")
out.append("GO")
out.append("")

# 12. ChuyenThamQuan - group visits by (factory, date)
out.append("-- ============ 12. ChuyenThamQuan ============")
out.append("SET IDENTITY_INSERT ChuyenThamQuan ON;")
trip_key_map = {}  # (factory, date) -> id
trip_id = 1
for v in visits:
    f = v['factory']
    d = v['date']
    if not d: continue
    key = (f, d)
    if key not in trip_key_map:
        nmid = nm_id_map.get(f)
        if not nmid: continue
        hinh_thuc = 'TrucTiep' if v['form'] == 'TrucTiep' else 'TrucTuyen'
        out.append(f"INSERT INTO ChuyenThamQuan (id, nha_may_id, lich_kien_tap_id, ngay_tham_quan, gio_bat_dau, gio_ket_thuc, hinh_thuc, suc_chua, trang_thai)")
        out.append(f"  VALUES ({trip_id}, {nmid}, 1, '{d}', '07:00', '17:00', N'{hinh_thuc}', 50, N'DaDienRa');")
        trip_key_map[key] = trip_id
        trip_id += 1

out.append("SET IDENTITY_INSERT ChuyenThamQuan OFF;")
out.append("GO")
out.append("")

# 13. PhieuDangKy
out.append("-- ============ 13. PhieuDangKy ============")
out.append("SET IDENTITY_INSERT PhieuDangKy ON;")
pdk_id = 1
pdk_map = {}  # (mssv, trip_id) -> pdk_id
for v in visits:
    mssv = v['mssv']
    f = v['factory']
    d = v['date']
    if not d: continue
    key = (f, d)
    tid = trip_key_map.get(key)
    sid = sv_id_map.get(mssv)
    if not tid or not sid: continue
    pdk_key = (mssv, tid)
    if pdk_key in pdk_map: continue
    out.append(f"INSERT INTO PhieuDangKy (id, sinh_vien_id, chuyen_tham_quan_id, trang_thai)")
    out.append(f"  VALUES ({pdk_id}, {sid}, {tid}, N'HoanThanh');")
    pdk_map[pdk_key] = pdk_id
    v['pdk_id'] = pdk_id
    pdk_id += 1

out.append("SET IDENTITY_INSERT PhieuDangKy OFF;")
out.append("GO")
out.append("")

# 14. DiemPhieuDangKy
out.append("-- ============ 14. DiemPhieuDangKy ============")
out.append("SET IDENTITY_INSERT DiemPhieuDangKy ON;")
dpdk_id = 1
for v in visits:
    mssv = v['mssv']
    f = v['factory']
    d = v['date']
    if not d: continue
    key = (f, d)
    tid = trip_key_map.get(key)
    if not tid: continue
    pdk_key = (mssv, tid)
    pid = pdk_map.get(pdk_key)
    if not pid: continue
    sc = v['score']
    bn = v['bonus']
    if sc is None: continue
    bn_val = bn if bn else 0
    out.append(f"INSERT INTO DiemPhieuDangKy (id, phieu_dang_ky_id, diem_chuan_bi, diem_cong)")
    out.append(f"  VALUES ({dpdk_id}, {pid}, {sc}, {bn_val});")
    dpdk_id += 1

out.append("SET IDENTITY_INSERT DiemPhieuDangKy OFF;")
out.append("GO")
out.append("")

out.append("PRINT N'Import hoàn tất.';")
out.append("GO")

# Write output
outpath = os.path.join(DB, 'QLKienTap_ImportData.sql')
with open(outpath, 'w', encoding='utf-8-sig') as f:
    f.write('\n'.join(out))

print(f"Generated {outpath}")
print(f"Students: {len(students)}, Teachers: {len(teachers)}, Factories: {len(factories)}")
print(f"Visits: {len(visits)}, Trips: {len(trip_key_map)}, Registrations: {len(pdk_map)}")
