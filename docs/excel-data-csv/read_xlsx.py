import openpyxl
import os
import sys
import io

# Force UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

DB_DIR = r'd:\DoAnTotNghiepCuNhan\HeThongQuanLyKienTap\DB'
OUTPUT = os.path.join(DB_DIR, 'xlsx_data_dump.txt')

files = [
    'DS TQNM ACECOOK T5-2025.xlsx',
    'DSSV TQNM Heineken 22052025.xlsx',
    u'D\u1eef li\u1ec7u \u0111i\u1ec3m qu\u00e1 tr\u00ecnh TQNM 2024-2025.xlsx',
    u'D\u1eef li\u1ec7u Kien tap 24-25 (G\u1eedi SV ki\u1ec3m tra).xlsx',
    'Mau chuan 15.3.26.xlsx',
]

with open(OUTPUT, 'w', encoding='utf-8') as out:
    for fname in files:
        fpath = os.path.join(DB_DIR, fname)
        out.write(f"\n{'='*100}\n")
        out.write(f"FILE: {fname}\n")
        out.write(f"{'='*100}\n")
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                out.write(f"\n--- Sheet: '{sheet_name}' ---\n")
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    row_count += 1
                    # Print all rows for complete data extraction
                    if row_count <= 200:  # Limit to first 200 rows per sheet
                        out.write(f"  R{row_count}: {row}\n")
                out.write(f"  TOTAL ROWS: {row_count}\n")
            wb.close()
        except Exception as e:
            out.write(f"  ERROR: {e}\n")

print("Done. Output written to xlsx_data_dump.txt")
